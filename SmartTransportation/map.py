import math

import folium
import openpyxl
import datetime
import numpy as np
import requests


def GetOrder(file, sheet):
    wb = openpyxl.load_workbook(file)
    Main = wb[sheet]
    Rows = Main.max_row
    GetBack = []
    MAX = 0
    for i in range(Rows - 1):
        GetBack.append([int(Main["C" + str(i + 2)].value), [format(float(Main["G" + str(i + 2)].value), '.6f'),
                                                            format(float(Main["F" + str(i + 2)].value), '.6f')]])
        if int(Main["C" + str(i + 2)].value) > MAX:
            MAX = int(Main["C" + str(i + 2)].value)
    return GetBack, MAX


def GetDataX(file, sheet, column):
    wb = openpyxl.load_workbook(file)
    Main = wb[sheet]
    Rows = Main.max_row
    Vec_1 = []
    for j in range(Rows - 1):
        Pos_1 = format(float(Main["G" + str(j + 2)].value), '.6f')
        Pos_2 = format(float(Main["F" + str(j + 2)].value), '.6f')
        x = (float(Pos_1) - 24.4186) / 0.01
        y = (float(Pos_2) - 117.9626) / 0.01
        f = '%Y-%m-%d %H:%M:%S'
        date = datetime.datetime.strptime(Main["D" + str(j + 2)].value, f)
        # weekday = datetime.date(int(date.strftime('%Y')), int(date.strftime('%m')),int(date.strftime(
        # '%d'))).isoweekday()
        Day = date.strftime('%Y') + '-' + date.strftime('%m') + '-' + date.strftime('%d')
        f_2 = '%H:%M:%S'
        base = datetime.datetime.strptime('6:00:00', f_2)  # 基准时间
        target = datetime.datetime.strptime(
            str(date.strftime('%H')) + ':' + str(date.strftime('%M')) + ':' + str(date.strftime('%S')), f_2)
        number = math.ceil(int((target - base).seconds) / 3600)
        pos = math.ceil(x) + (math.ceil(y) - 1) * column
        if pos < 0:
            continue
        Vec_1.append([pos, Day, number])
    return Vec_1


def GetDataY(Vec):
    Vec_2 = []
    Vec_3 = []
    for v in Vec:
        if v not in Vec_2:
            Vec_2.append(v)
            Vec_3.append(1)
        else:
            Vec_3[len(Vec_3) - 1] += 1
    f = "%Y-%m-%d"
    i = 0
    for v in Vec_2:
        date = datetime.datetime.strptime(v[1], f)
        weekday = datetime.date(int(date.strftime('%Y')), int(date.strftime('%m')),
                                int(date.strftime('%d'))).isoweekday()
        Vec_2[i][1] = weekday
        i += 1
    return np.asarray(Vec_2), np.asarray(Vec_3)


def GetPos(date, file, sheet, number, scale):
    limit = number * scale + 6
    f = '%Y-%m-%d %H:%M:%S'
    MinTime = datetime.datetime.strptime(date + ' ' + str(limit - scale) + ':00:00', f)
    MaxTime = datetime.datetime.strptime(date + ' ' + str(limit) + ':00:00', f)
    # print(MaxTime)
    wb = openpyxl.load_workbook(file)
    Main = wb[sheet]
    Rows = Main.max_row
    Pos_1 = []
    Pos_2 = []
    for j in range(Rows - 1):
        if datetime.datetime.strptime(Main["M" + str(j + 2)].value, f) >= MinTime and datetime.datetime.strptime(
                Main["M" + str(j + 2)].value, f) < MaxTime:
            Pos_1.append(format(float(Main["G" + str(j + 2)].value), '.6f'))
            Pos_2.append(format(float(Main["F" + str(j + 2)].value), '.6f'))
    return Pos_1, Pos_2




if __name__ == '__main__':
    '''filename = "数据集.xlsx"
    pos_1, pos_2 = GetPos('2023-02-03', filename, 'default')
    start_pos = []
    i = 0
    for p in pos_1:
        start_pos.append([float(p), float(pos_2[i])])
        i += 1
    print(start_pos)
    m = folium.Map([24.484514, 118.14519],
                   tiles='http://webrd02.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=7&x={x}&y={y}&z={z}',
                   attr='高德-中英文对照',
                   zoom_start=15,
                   )
    for sp in start_pos:
        folium.Marker(
            location=sp,
            popup='Mt. Hood Meadows',
            icon=folium.Icon(icon='cloud')
        ).add_to(m)
    # 输出当天数据量大小
    # print(len(start_pos))
    m.save('map.html')'''
    filename = "数据集.xlsx"
    print(GetDataY(GetDataX(filename, 'default', 24)))
